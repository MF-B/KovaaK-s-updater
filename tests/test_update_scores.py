import csv
import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

import update_scores


def create_sample_db(tmp_path: Path) -> Path:
    db_path = tmp_path / "stats.db"
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute(
        "CREATE TABLE scenario_scores (id INTEGER PRIMARY KEY, scenarioName TEXT, mode TEXT, totalScore REAL)"
    )
    cursor.executemany(
        "INSERT INTO scenario_scores (scenarioName, mode, totalScore) VALUES (?, ?, ?)",
        [
            ("Air Angelic 4", "test", 1500),
            ("Air Angelic 4", "test", 1700),
            ("Whisphere 80%", "test", 6300),
        ],
    )
    conn.commit()
    conn.close()
    return db_path


def create_sample_workbook(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Scenarios"
    ws.append(["Viscose Test", None, None, None, None])
    ws.append([None, None, "Scenario", None, "High Score"])
    ws.append([None, None, "Air Angelic 4", None, None])
    ws.append([None, None, "Whisphere 80%", None, None])
    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)
    return file_path


class UpdateScoresTestCase(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self.tmp_dir.name)

    def tearDown(self):
        self.tmp_dir.cleanup()

    def test_fetch_scores_and_update_workbook(self):
        db_path = create_sample_db(self.tmp_path)
        excel_path = create_sample_workbook(self.tmp_path)

        scores, selection = update_scores.fetch_scores_from_db(
            db_path,
            table="scenario_scores",
            name_column="scenarioName",
            score_column="totalScore",
        )

        self.assertEqual(selection.table, "scenario_scores")
        self.assertEqual(scores[update_scores.normalize_key("Air Angelic 4")], 1700)

        stats, workbook = update_scores.update_workbook(
            excel_path,
            scores,
            sheets=["Test Scenarios"],
        )

        self.assertEqual(stats[0].updated, 2)
        ws = workbook["Test Scenarios"]
        self.assertEqual(ws.cell(row=3, column=5).value, 1700)
        self.assertEqual(ws.cell(row=4, column=5).value, 6300)

    def test_fetch_scores_from_csv(self):
        csv_file = self.tmp_path / "scores.csv"
        with csv_file.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Scenario Name", "Score", "Shots"])
            writer.writerow(["Air Angelic 4", "1500", "10"])
            writer.writerow(["Air Angelic 4", "1700", "10"])
            writer.writerow(["Whisphere 80%", "6300", "20"])

        scores = update_scores.fetch_scores_from_csv([csv_file])
        self.assertEqual(scores[update_scores.normalize_key("Air Angelic 4")], 1700)
        self.assertEqual(scores[update_scores.normalize_key("Whisphere 80%")], 6300)

    def test_fetch_scores_from_key_value_csv(self):
        csv_file = self.tmp_path / "kv.csv"
        csv_file.write_text(
            "Score:,3180.0\nMBS Points:,0\nScenario:,Whisphere 80%\n",
            encoding="utf-8",
        )

        scores = update_scores.fetch_scores_from_csv([csv_file])
        self.assertEqual(scores[update_scores.normalize_key("Whisphere 80%")], 3180.0)

    def test_csv_with_table_but_no_scenario_column(self):
        csv_file = self.tmp_path / "table_meta.csv"
        csv_file.write_text(
            "Kill #,Shots,Hits\n"
            "1,10,8\n"
            "2,12,9\n\n"
            "Score:,1234.5\n"
            "Scenario:,Metadata Map\n",
            encoding="utf-8",
        )

        scores = update_scores.fetch_scores_from_csv([csv_file])
        key = update_scores.normalize_key("Metadata Map")
        self.assertEqual(scores[key], 1234.5)

    def test_update_sheet_with_formula_headers(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Formula Sheet"
        ws.cell(row=1, column=3, value='=" Scenario"')
        ws.cell(row=1, column=5, value='=" High Score"')
        ws.cell(row=2, column=3, value="Whisphere 80%")
        ws.cell(row=3, column=3, value="Air Angelic 4")

        stats = update_scores.update_sheet(
            ws,
            {
                update_scores.normalize_key("Whisphere 80%"): 4000,
                update_scores.normalize_key("Air Angelic 4"): 1700,
            },
        )

        self.assertEqual(stats.updated, 2)
        self.assertEqual(ws.cell(row=2, column=5).value, 4000)
        self.assertEqual(ws.cell(row=3, column=5).value, 1700)

    def test_update_sheet_with_formula_scenario_values(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Formula Scenario"
        ws.cell(row=1, column=3, value="Scenario")
        ws.cell(row=1, column=5, value="High Score")
        ws.cell(row=2, column=3, value='=IFERROR(__xludf.DummyFunction("""source"""),"Whisphere 80%")')
        ws.cell(row=3, column=3, value='=IFERROR(__xludf.DummyFunction("""source"""),"Controlsphere rAim Easy 90%")')

        stats = update_scores.update_sheet(
            ws,
            {
                update_scores.normalize_key("Whisphere 80%"   ): 3180,
                update_scores.normalize_key("Controlsphere rAim Easy 90%"   ): 3261,
            },
        )

        self.assertEqual(stats.updated, 2)
        self.assertEqual(ws.cell(row=2, column=5).value, 3180)
        self.assertEqual(ws.cell(row=3, column=5).value, 3261)

    def test_update_sheet_with_merged_high_score_columns(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged High Score"
        ws.cell(row=1, column=3, value="Scenario")
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
        ws.cell(row=1, column=5, value="High Score")
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 25
        ws.cell(row=2, column=3, value="Whisphere 80%")
        ws.cell(row=3, column=3, value="Controlsphere rAim Easy 90%")

        stats = update_scores.update_sheet(
            ws,
            {
                update_scores.normalize_key("Whisphere 80%"   ): 3180,
                update_scores.normalize_key("Controlsphere rAim Easy 90%"   ): 3261,
            },
        )

        self.assertEqual(stats.updated, 2)
        self.assertIsNone(ws.cell(row=2, column=5).value)
        self.assertEqual(ws.cell(row=2, column=6).value, 3180)
        self.assertIsNone(ws.cell(row=3, column=5).value)
        self.assertEqual(ws.cell(row=3, column=6).value, 3261)

    def test_update_sheet_prefers_rightmost_when_equal_width(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Equal Width"
        ws.cell(row=1, column=3, value="Scenario")
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
        ws.cell(row=1, column=5, value="High Score")
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 20
        ws.cell(row=2, column=3, value="Scenario A")
        ws.cell(row=3, column=3, value="Scenario B")
        ws.cell(row=3, column=5, value=999)

        stats = update_scores.update_sheet(
            ws,
            {
                update_scores.normalize_key("Scenario A"): 111,
                update_scores.normalize_key("Scenario B"): 222,
            },
        )

        self.assertEqual(stats.updated, 2)
        self.assertEqual(ws.cell(row=2, column=6).value, 111)
        # 第二行的列 E 已有值，因此应当 fallback 到列 F
        self.assertEqual(ws.cell(row=3, column=6).value, 222)


if __name__ == "__main__":
    unittest.main()
