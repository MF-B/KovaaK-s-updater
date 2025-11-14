#!/usr/bin/env python3
"""Utility for syncing KovaaK stats.db high scores into the Viscose Excel tracker.

Phase 1 deliverable: a manually executed script that reads KovaaK's SQLite database,
extracts per-scenario high scores, and writes them into the "High Score" column
inside the provided workbook.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import logging
import re
import sqlite3
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

try:
    import pandas as pd  # Optional, used for pretty console summaries.
except ImportError:  # pragma: no cover - pandas is optional but recommended.
    pd = None  # type: ignore

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

LOGGER = logging.getLogger("kovaak_updater")
NAME_KEYWORDS = ["scenario", "drill", "map", "task", "playlist", "name", "title"]
SCORE_KEYWORDS = ["score", "result", "value", "points"]
NUMERIC_HINTS = ["int", "real", "num", "double", "float"]
DEFAULT_SCENARIO_SHEET_FRAGMENT = "scenario"
HOME = Path.home()
DEFAULT_STATS_DIRS = [
    Path("C:/Program Files (x86)/Steam/steamapps/common/KovaaKs FPS Aim Trainer/stats"),
    Path("C:/Program Files/Steam/steamapps/common/KovaaKs FPS Aim Trainer/stats"),
    HOME / ".steam/steam/steamapps/common/KovaaKs FPS Aim Trainer/stats",
    HOME / ".local/share/Steam/steamapps/common/KovaaKs FPS Aim Trainer/stats",
    HOME / "Steam/steamapps/common/KovaaKs FPS Aim Trainer/stats",
    HOME / ".steam/steam/steamapps/common/FPSAimTrainer/FPSAimTrainer/stats",
]


def try_parse_float(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None
    cleaned = str(value).replace(",", "").strip()
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


@dataclass
class TableSelection:
    table: str
    name_column: str
    score_column: str


@dataclass
class SheetHeader:
    header_row: int
    scenario_col: int
    high_score_cols: List[int]

    @property
    def primary_high_score_col(self) -> int:
        return self.high_score_cols[0]


@dataclass
class SheetUpdateStats:
    sheet: str
    updated: int
    missing: List[str]


@dataclass
class UpdateSettings:
    excel: Path = Path("Viscose Benchmarks Beta.xlsx")
    stats_db: Optional[Path] = None
    csv_paths: List[Path] = field(default_factory=list)
    csv_pattern: str = "*.csv"
    csv_delimiter: str = ","
    sheets: Optional[List[str]] = None
    table: Optional[str] = None
    name_column: Optional[str] = None
    score_column: Optional[str] = None
    output: Optional[Path] = None
    dry_run: bool = False


@dataclass
class WatchSettings:
    paths: List[Path] = field(default_factory=list)
    debounce_seconds: float = 2.0
    run_on_start: bool = True


def normalize_key(value: str) -> str:
    clean = " ".join(value.split())
    return clean.strip().lower()


QUOTED_TEXT_PATTERN = re.compile(r'"([^\"]+)"')
DEFAULT_EXCEL_COLUMN_WIDTH = 8.43


def extract_header_text(value) -> Optional[str]:
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    if text.startswith("="):
        match = re.match(r'^="([^"]+)"$', text)
        if match:
            text = match.group(1)
        else:
            quoted = QUOTED_TEXT_PATTERN.search(text)
            if quoted:
                text = quoted.group(1)
            else:
                text = text.lstrip("=")
    text = text.strip()
    return text.lower() if text else None


def extract_cell_text(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        text = str(value)
    else:
        text = str(value).strip()
    if not text:
        return None
    if text.startswith("="):
        literals: List[str] = []
        current: List[str] = []
        in_literal = False
        i = 0
        while i < len(text):
            char = text[i]
            if char == '"':
                if in_literal:
                    next_char = text[i + 1] if i + 1 < len(text) else None
                    if next_char == '"':
                        current.append('"')
                        i += 2
                        continue
                    literals.append("".join(current))
                    current = []
                    in_literal = False
                else:
                    in_literal = True
                i += 1
                continue
            if in_literal:
                current.append(char)
            i += 1
        if current and in_literal:
            literals.append("".join(current))
        if literals:
            text = literals[-1]
        else:
            text = text.lstrip("=")
    text = text.strip()
    return text or None


def resolve_high_score_columns(ws: Worksheet, header_row: int, base_col: int) -> List[int]:
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= header_row <= merged_range.max_row
            and merged_range.min_col <= base_col <= merged_range.max_col
        ):
            return list(range(merged_range.min_col, merged_range.max_col + 1))
    return [base_col]


def choose_high_score_column(ws: Worksheet, candidate_cols: Sequence[int], row_idx: int) -> Optional[int]:
    best_col = None
    best_width = -1.0
    for col in candidate_cols:
        cell = ws.cell(row=row_idx, column=col)
        value = cell.value
        if value not in (None, ""):
            continue
        letter = get_column_letter(col)
        dimension = ws.column_dimensions.get(letter)
        width = dimension.width if dimension and dimension.width is not None else DEFAULT_EXCEL_COLUMN_WIDTH
        if width > best_width or (width == best_width and (best_col is None or col > best_col)):
            best_col = col
            best_width = width
    if best_col is not None:
        return best_col
    return max(candidate_cols) if candidate_cols else None


def quote_identifier(identifier: str) -> str:
    escaped = identifier.replace('"', '""')
    return f'"{escaped}"'


def list_existing_stats_dirs() -> List[Path]:
    return [path for path in DEFAULT_STATS_DIRS if path.exists()]


def resolve_config_path(value: str, base_dir: Path) -> Path:
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = (base_dir / path).resolve()
    return path


def ensure_list(value: Any) -> List[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def apply_update_settings_from_dict(settings: UpdateSettings, data: Dict[str, Any], base_dir: Path) -> None:
    if "excel" in data:
        settings.excel = resolve_config_path(data["excel"], base_dir)
    if "stats_db" in data:
        stats_value = data["stats_db"]
        settings.stats_db = resolve_config_path(stats_value, base_dir) if stats_value else None
    if "csv_paths" in data:
        settings.csv_paths = [resolve_config_path(item, base_dir) for item in ensure_list(data["csv_paths"])]
    if "csv_pattern" in data:
        settings.csv_pattern = str(data["csv_pattern"])
    if "csv_delimiter" in data:
        settings.csv_delimiter = str(data["csv_delimiter"])
    if "sheets" in data:
        settings.sheets = [str(item) for item in ensure_list(data["sheets"])]
    if "table" in data:
        settings.table = data["table"] or None
    if "name_column" in data:
        settings.name_column = data["name_column"] or None
    if "score_column" in data:
        settings.score_column = data["score_column"] or None
    if "output" in data:
        output_value = data["output"]
        settings.output = resolve_config_path(output_value, base_dir) if output_value else None
    if "dry_run" in data:
        settings.dry_run = bool(data["dry_run"])


def apply_watch_settings_from_dict(watch: WatchSettings, data: Dict[str, Any], base_dir: Path) -> None:
    if "paths" in data:
        watch.paths = [resolve_config_path(item, base_dir) for item in ensure_list(data["paths"])]
    if "debounce_seconds" in data:
        watch.debounce_seconds = float(data["debounce_seconds"])
    if "run_on_start" in data:
        watch.run_on_start = bool(data["run_on_start"])


def discover_config_path(cli_path: Optional[Path]) -> Optional[Path]:
    if cli_path:
        return cli_path.expanduser().resolve()
    default_path = Path("config.json")
    if default_path.exists():
        return default_path.resolve()
    return None


def load_settings_from_config(
    config_path: Optional[Path],
) -> Tuple[Optional[UpdateSettings], Optional[WatchSettings]]:
    if config_path is None:
        return None, None
    config_path = config_path.expanduser().resolve()
    if not config_path.exists():
        raise FileNotFoundError(f"未找到配置文件: {config_path}")

    with config_path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)

    base_dir = config_path.parent
    settings = UpdateSettings()
    apply_update_settings_from_dict(settings, data, base_dir)
    watch_settings = None
    if isinstance(data.get("watch"), dict):
        watch_settings = WatchSettings()
        apply_watch_settings_from_dict(watch_settings, data["watch"], base_dir)
    return settings, watch_settings


def merge_settings(base_settings: Optional[UpdateSettings], args: argparse.Namespace) -> UpdateSettings:
    settings = UpdateSettings()
    if base_settings:
        settings = UpdateSettings(
            excel=base_settings.excel,
            stats_db=base_settings.stats_db,
            csv_paths=list(base_settings.csv_paths),
            csv_pattern=base_settings.csv_pattern,
            csv_delimiter=base_settings.csv_delimiter,
            sheets=list(base_settings.sheets) if base_settings.sheets else None,
            table=base_settings.table,
            name_column=base_settings.name_column,
            score_column=base_settings.score_column,
            output=base_settings.output,
            dry_run=base_settings.dry_run,
        )

    def _maybe_path(value: Optional[Path]) -> Optional[Path]:
        if value is None:
            return None
        return value.expanduser().resolve()

    if args.excel:
        settings.excel = _maybe_path(args.excel) or settings.excel
    if args.db:
        settings.stats_db = _maybe_path(args.db)
    if args.csv:
        for csv_entry in args.csv:
            resolved_csv = _maybe_path(csv_entry)
            if resolved_csv:
                settings.csv_paths.append(resolved_csv)
    if args.csv_pattern:
        settings.csv_pattern = args.csv_pattern
    if args.csv_delimiter:
        settings.csv_delimiter = args.csv_delimiter
    if args.sheets:
        settings.sheets = args.sheets
    if args.table:
        settings.table = args.table
    if args.name_column:
        settings.name_column = args.name_column
    if args.score_column:
        settings.score_column = args.score_column
    if args.output:
        settings.output = _maybe_path(args.output)
    if args.dry_run:
        settings.dry_run = True

    if not settings.csv_paths:
        settings.csv_paths = []

    return settings


def prepare_watch_settings(
    update_settings: UpdateSettings,
    watch_settings: Optional[WatchSettings],
) -> WatchSettings:
    settings = WatchSettings()
    if watch_settings:
        settings.paths = [path.expanduser().resolve() for path in watch_settings.paths]
        settings.debounce_seconds = watch_settings.debounce_seconds
        settings.run_on_start = watch_settings.run_on_start

    if not settings.paths:
        derived_paths: List[Path] = []
        if update_settings.stats_db:
            derived_paths.append(update_settings.stats_db)
        derived_paths.extend(update_settings.csv_paths)
        if not derived_paths:
            derived_paths.extend(list_existing_stats_dirs())
        settings.paths = [path.expanduser().resolve() for path in derived_paths if path]

    return settings


def detect_table(conn: sqlite3.Connection) -> Optional[TableSelection]:
    cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
    best_score = -1.0
    best_selection: Optional[TableSelection] = None

    for (table_name,) in cursor.fetchall():
        table_info = conn.execute(f"PRAGMA table_info({quote_identifier(table_name)})").fetchall()
        if not table_info:
            continue

        name_candidates = []
        score_candidates = []
        for column in table_info:
            col_name = column[1]
            col_type = (column[2] or '').lower()
            lower_name = col_name.lower()
            name_weight = max((len(kw) for kw in NAME_KEYWORDS if kw in lower_name), default=0)
            score_weight = max((len(kw) for kw in SCORE_KEYWORDS if kw in lower_name), default=0)
            numeric_hint = any(hint in col_type for hint in NUMERIC_HINTS)

            if name_weight:
                name_candidates.append((col_name, name_weight))
            if score_weight or numeric_hint:
                bonus = 2 if score_weight else 1
                score_candidates.append((col_name, score_weight + bonus))

        if not name_candidates or not score_candidates:
            continue

        table_weight = (max(n[1] for n in name_candidates) + max(s[1] for s in score_candidates))
        if table_weight > best_score:
            best_score = table_weight
            best_selection = TableSelection(
                table=table_name,
                name_column=max(name_candidates, key=lambda x: x[1])[0],
                score_column=max(score_candidates, key=lambda x: x[1])[0],
            )

    return best_selection


def fetch_scores_from_db(
    db_path: Path,
    table: Optional[str] = None,
    name_column: Optional[str] = None,
    score_column: Optional[str] = None,
) -> Tuple[Dict[str, float], TableSelection]:
    conn = sqlite3.connect(db_path)
    try:
        conn.row_factory = sqlite3.Row
        selection = TableSelection(table, name_column, score_column) if table and name_column and score_column else None
        if selection is None:
            selection = detect_table(conn)
        if selection is None:
            raise RuntimeError(
                "无法根据 stats.db 自动识别场景得分表。请使用 --table/--name-column/--score-column 参数指定。"
            )

        LOGGER.info(
            "使用数据表 %s (名称列=%s, 分数列=%s)",
            selection.table,
            selection.name_column,
            selection.score_column,
        )
        name_sql = quote_identifier(selection.name_column)
        score_sql = quote_identifier(selection.score_column)
        table_sql = quote_identifier(selection.table)
        query = f"""
            SELECT {name_sql} AS scenario_name,
                   MAX(CAST({score_sql} AS REAL)) AS high_score
            FROM {table_sql}
            WHERE {name_sql} IS NOT NULL AND TRIM({name_sql}) <> ''
              AND {score_sql} IS NOT NULL
            GROUP BY {name_sql}
        """
        result: Dict[str, float] = {}
        for row in conn.execute(query):
            name_val = row["scenario_name"]
            score_val = row["high_score"]
            if name_val is None or score_val is None:
                continue
            normalized = normalize_key(str(name_val))
            result[normalized] = float(score_val)

        if not result:
            raise RuntimeError("无法从 stats.db 获取任何分数，请确认数据库内包含游戏数据。")

        return result, selection
    finally:
        conn.close()


def detect_csv_columns(headers: Sequence[Optional[str]]) -> Tuple[Optional[str], Optional[str]]:
    scenario_col = None
    scenario_weight = 0
    score_col = None
    score_weight = 0
    for header in headers:
        if not header:
            continue
        text = header.strip().lower()
        n_weight = max((len(kw) for kw in NAME_KEYWORDS if kw in text), default=0)
        s_weight = max((len(kw) for kw in SCORE_KEYWORDS if kw in text), default=0)
        if n_weight > scenario_weight:
            scenario_col = header
            scenario_weight = n_weight
        if s_weight > score_weight:
            score_col = header
            score_weight = s_weight
    return scenario_col, score_col


def select_numeric_column(rows: List[Dict[str, str]], headers: Sequence[Optional[str]]) -> Optional[str]:
    for header in headers:
        if not header:
            continue
        numeric_values = 0
        for row in rows:
            value = row.get(header)
            parsed = try_parse_float(value) if value is not None else None
            if parsed is None:
                numeric_values = 0
                break
            numeric_values += 1
        if numeric_values:
            return header
    return None

def extract_key_value_score(rows: Sequence[Sequence[str]], fallback_name: str) -> Optional[Tuple[str, float]]:
    scenario_name: Optional[str] = None
    best_score: Optional[float] = None

    for row in rows:
        trimmed = [cell.strip() for cell in row if cell and cell.strip()]
        if not trimmed:
            continue
        idx = 0
        while idx < len(trimmed):
            cell = trimmed[idx]
            key = None
            value = None
            if ":" in cell:
                key_part, _, remainder = cell.partition(":")
                key = key_part.strip()
                value = remainder.strip()
                if not value and idx + 1 < len(trimmed):
                    value = trimmed[idx + 1]
                    idx += 1
            elif cell.endswith(":") and idx + 1 < len(trimmed):
                key = cell.rstrip(":").strip()
                value = trimmed[idx + 1]
                idx += 1
            else:
                idx += 1
                continue

            idx += 1
            if not key:
                continue

            key_lower = key.lower()
            normalized_value = value.strip() if value else ""
            if key_lower == "scenario" or any(kw in key_lower for kw in NAME_KEYWORDS):
                scenario_name = normalized_value or scenario_name
            if "score" in key_lower:
                parsed = try_parse_float(normalized_value)
                if parsed is not None and (best_score is None or parsed > best_score):
                    best_score = parsed

    if best_score is None:
        return None

    scenario = scenario_name or fallback_name
    return scenario, best_score


def fetch_scores_from_csv(
    csv_files: Sequence[Path],
    delimiter: str = ",",
) -> Dict[str, float]:
    if not csv_files:
        raise RuntimeError("未提供任何 CSV 文件。")

    aggregated: Dict[str, float] = {}
    for csv_path in csv_files:
        if not csv_path.exists():
            LOGGER.warning("CSV 文件不存在: %s", csv_path)
            continue
        raw_text = csv_path.read_text(encoding="utf-8-sig")
        dict_reader = csv.DictReader(io.StringIO(raw_text), delimiter=delimiter)
        rows = [row for row in dict_reader]
        headers = dict_reader.fieldnames or []
        plain_rows = list(csv.reader(io.StringIO(raw_text), delimiter=delimiter))

        fallback_name = csv_path.stem
        metadata_result = extract_key_value_score(plain_rows, fallback_name)
        metadata_present = metadata_result is not None
        updated_rows = 0

        scenario_col, score_col = detect_csv_columns(headers)
        if score_col is None and rows:
            score_col = select_numeric_column(rows, headers)

        can_use_table = scenario_col is not None and score_col is not None and rows
        if can_use_table:
            for row in rows:
                scenario_value = row.get(scenario_col)
                if scenario_value is None:
                    continue
                scenario_name = str(scenario_value).strip()
                if not scenario_name:
                    continue

                raw_score = row.get(score_col)
                score_value = try_parse_float(raw_score) if raw_score is not None else None
                if score_value is None:
                    continue

                key = normalize_key(scenario_name)
                previous = aggregated.get(key)
                if previous is None or score_value > previous:
                    aggregated[key] = score_value
                    updated_rows += 1
        elif scenario_col is not None and rows:
            LOGGER.debug("CSV %s 检测到场景列但缺少分数字段，已跳过表格数据。", csv_path.name)

        if metadata_result:
            scenario_name, score_value = metadata_result
            key = normalize_key(scenario_name)
            previous = aggregated.get(key)
            if previous is None or score_value > previous:
                aggregated[key] = score_value
                updated_rows += 1

        if updated_rows or metadata_present:
            extra = " + 元数据" if metadata_present else ""
            LOGGER.debug("解析 CSV %s -> 记录 %d 条%s", csv_path.name, updated_rows, extra)
        else:
            LOGGER.warning("CSV %s 未找到有效的场景/分数记录，已跳过。", csv_path)

    if not aggregated:
        raise RuntimeError("CSV 文件中未找到可用的场景分数。")

    return aggregated


def detect_header(ws: Worksheet) -> Optional[SheetHeader]:
    max_row = min(ws.max_row, 30)
    scenario_col = None
    high_score_col = None
    header_row = None

    for row in range(1, max_row + 1):
        for cell in ws[row]:
            value = cell.value
            text = extract_header_text(value)
            if text is None:
                continue
            if text == "scenario":
                scenario_col = cell.column
                header_row = cell.row
            elif text == "high score":
                high_score_col = cell.column
                header_row = header_row or cell.row
        if scenario_col and high_score_col:
            high_score_cols = resolve_high_score_columns(ws, header_row, high_score_col)
            return SheetHeader(header_row=header_row, scenario_col=scenario_col, high_score_cols=high_score_cols)

    return None


def update_sheet(ws: Worksheet, scores: Dict[str, float]) -> SheetUpdateStats:
    header = detect_header(ws)
    if header is None:
        LOGGER.warning("工作表 %s 未找到 'Scenario'/'High Score' 标题，已跳过。", ws.title)
        return SheetUpdateStats(sheet=ws.title, updated=0, missing=[])

    updated = 0
    missing: List[str] = []
    for row_idx in range(header.header_row + 1, ws.max_row + 1):
        scenario_cell = ws.cell(row=row_idx, column=header.scenario_col)
        scenario_name = extract_cell_text(scenario_cell.value)
        if not scenario_name:
            continue
        normalized = normalize_key(scenario_name)
        if normalized not in scores:
            missing.append(scenario_name)
            continue
        score_value = scores[normalized]
        target_column = choose_high_score_column(ws, header.high_score_cols, row_idx)
        if target_column is None:
            LOGGER.warning("工作表 %s 找不到可写入的 High Score 列，已跳过行 %d。", ws.title, row_idx)
            continue
        target_cell = ws.cell(row=row_idx, column=target_column)
        target_cell.value = round(score_value, 2)
        updated += 1

    return SheetUpdateStats(sheet=ws.title, updated=updated, missing=missing)


def select_default_sheets(sheet_names: Sequence[str]) -> List[str]:
    candidates = [name for name in sheet_names if DEFAULT_SCENARIO_SHEET_FRAGMENT in name.lower()]
    return candidates or list(sheet_names)


def collect_csv_files(entries: Sequence[Path], pattern: str = "*.csv") -> List[Path]:
    files: List[Path] = []
    for entry in entries:
        resolved = entry.expanduser().resolve()
        if resolved.is_file() and resolved.suffix.lower() == ".csv":
            files.append(resolved)
        elif resolved.is_dir():
            files.extend(sorted(resolved.glob(pattern)))
        else:
            LOGGER.warning("未找到 CSV 路径: %s", resolved)
    return files


def guess_default_csv_files(pattern: str = "*.csv") -> List[Path]:
    files: List[Path] = []
    for directory in DEFAULT_STATS_DIRS:
        stats_dir = directory.expanduser()
        if stats_dir.is_dir():
            files.extend(sorted(stats_dir.glob(pattern)))
    return files


def update_workbook(
    workbook_path: Path,
    scores: Dict[str, float],
    sheets: Optional[Sequence[str]] = None,
) -> Tuple[List[SheetUpdateStats], Workbook]:
    wb = load_workbook(workbook_path)
    target_sheets = sheets or select_default_sheets(wb.sheetnames)
    stats: List[SheetUpdateStats] = []
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            LOGGER.warning("工作表 %s 不存在，已跳过。", sheet_name)
            continue
        ws = wb[sheet_name]
        stats.append(update_sheet(ws, scores))

    return stats, wb


def guess_default_db_path() -> Optional[Path]:
    for stats_dir in DEFAULT_STATS_DIRS:
        candidate = stats_dir / "stats.db"
        if candidate.exists():
            return candidate
    return None


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="同步 KovaaK 高分到 Excel")
    parser.add_argument("--config", type=Path, default=None, help="配置文件路径 (JSON)")
    parser.add_argument("--db", type=Path, default=None, help="stats.db 路径 (默认自动探测常见路径)")
    parser.add_argument("--excel", type=Path, default=None, help="Excel 文件路径")
    parser.add_argument("--output", type=Path, default=None, help="保存为新文件的路径 (默认原地覆盖)")
    parser.add_argument("--sheets", nargs="*", default=None, help="需要更新的工作表名称 (默认自动选择包含 Scenarios 的工作表)")
    parser.add_argument("--csv", action="append", type=Path, help="CSV 文件或目录 (可重复指定)")
    parser.add_argument("--csv-pattern", default=None, help="当 --csv 指向目录时使用的 glob 模式")
    parser.add_argument("--csv-delimiter", default=None, help="CSV 文件使用的分隔符 (默认逗号)")
    parser.add_argument("--table", help="数据库中包含分数的表名")
    parser.add_argument("--name-column", dest="name_column", help="场景名称列")
    parser.add_argument("--score-column", dest="score_column", help="分数列")
    parser.add_argument("--dry-run", action="store_true", help="仅输出日志，不写回 Excel")
    parser.add_argument("--verbose", action="store_true", help="输出调试日志")
    return parser


def render_summary(stats: Iterable[SheetUpdateStats]):
    rows = [
        {
            "Sheet": item.sheet,
            "Updated": item.updated,
        }
        for item in stats
        if item.updated > 0
    ]

    if not rows:
        print("没有任何得分被写入 (0 条更新)。")
        return

    if pd is not None:
        df = pd.DataFrame(rows)
        print(df.to_string(index=False))
    else:
        for row in rows:
            print(f"工作表 {row['Sheet']} -> 更新 {row['Updated']} 条")


def run_update(settings: UpdateSettings) -> Tuple[List[SheetUpdateStats], Path]:
    excel_path = settings.excel.expanduser().resolve()
    if not excel_path.exists():
        raise FileNotFoundError(f"未找到 Excel 文件: {excel_path}")

    db_path = settings.stats_db.expanduser().resolve() if settings.stats_db else None
    if db_path and not db_path.exists():
        raise FileNotFoundError(f"数据库路径不存在: {db_path}")

    if db_path is None:
        guessed_db = guess_default_db_path()
        if guessed_db:
            db_path = guessed_db.expanduser().resolve()
            LOGGER.info("自动检测到 stats.db: %s", db_path)

    csv_pattern = settings.csv_pattern or "*.csv"
    csv_files: List[Path] = []
    if settings.csv_paths:
        csv_files = collect_csv_files(settings.csv_paths, pattern=csv_pattern)
    if not csv_files:
        csv_files = guess_default_csv_files(pattern=csv_pattern)

    data_source = None
    if db_path and db_path.exists():
        data_source = "db"
    elif csv_files:
        data_source = "csv"

    if data_source is None:
        raise RuntimeError("未能找到 stats.db，也没有可用的 CSV 文件。请在配置或命令行中提供路径。")

    if data_source == "db":
        scores, _selection = fetch_scores_from_db(
            db_path=db_path,
            table=settings.table,
            name_column=settings.name_column,
            score_column=settings.score_column,
        )
        LOGGER.info("共加载 %d 条场景分数记录 (来自 SQLite)", len(scores))
    else:
        delimiter = settings.csv_delimiter or ","
        scores = fetch_scores_from_csv(csv_files, delimiter=delimiter)
        LOGGER.info("共加载 %d 条场景分数记录 (来自 CSV)", len(scores))

    workbook_stats, workbook = update_workbook(
        workbook_path=excel_path,
        scores=scores,
        sheets=settings.sheets,
    )

    output_path = settings.output.expanduser().resolve() if settings.output else excel_path
    if settings.dry_run:
        LOGGER.info("Dry-run 模式，未对 Excel 进行写入 (目标: %s)", output_path)
    else:
        if output_path == excel_path:
            LOGGER.info("保存修改 -> %s", output_path)
        else:
            LOGGER.info("保存到新文件 -> %s", output_path)
        workbook.save(output_path)

    return workbook_stats, output_path


def main():
    parser = build_parser()
    args = parser.parse_args()
    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO, format="[%(levelname)s] %(message)s")
    config_path = discover_config_path(args.config)
    if config_path:
        LOGGER.info("使用配置文件: %s", config_path)
    base_settings, _watch = load_settings_from_config(config_path)
    settings = merge_settings(base_settings, args)

    if args.db and settings.stats_db and not settings.stats_db.exists():
        parser.error(f"数据库路径不存在: {settings.stats_db}")

    if args.csv:
        csv_files = collect_csv_files(settings.csv_paths, pattern=settings.csv_pattern)
        if not csv_files:
            parser.error("在 --csv 指定的路径中未找到任何 CSV 文件")

    try:
        workbook_stats, output_path = run_update(settings)
    except FileNotFoundError as exc:
        parser.error(str(exc))
    except RuntimeError as exc:
        LOGGER.error("更新失败: %s", exc)
        raise SystemExit(1)

    render_summary(workbook_stats)


if __name__ == "__main__":
    main()
